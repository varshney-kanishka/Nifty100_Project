import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

print("=" * 70)
print("DAY 20 - PEER COMPARISON REPORT")
print("=" * 70)

BASE_DIR = Path(__file__).resolve().parents[2]

DB = BASE_DIR / "data/database/nifty100.db"

conn = sqlite3.connect(DB)

ratios = pd.read_sql(
    "SELECT * FROM financial_ratios",
    conn,
)

peer = pd.read_sql(
    "SELECT * FROM peer_groups",
    conn,
)

companies = pd.read_sql(
    "SELECT id, company_name FROM companies",
    conn,
)

percentiles = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    conn,
)

print()

print("Tables Loaded")

print("Ratios :", len(ratios))
print("Peer :", len(peer))
print("Percentiles :", len(percentiles))

df = ratios.merge(
    peer,
    on="company_id",
    how="left",
)

df = df.merge(
    companies,
    left_on="company_id",
    right_on="id",
    how="left",
)

print()

print("Merged Successfully")

print(df.head())

output = BASE_DIR / "output"

output.mkdir(exist_ok=True)

excel_file = output / "peer_comparison.xlsx"

writer = pd.ExcelWriter(
    excel_file,
    engine="openpyxl",
)

groups = (
    df["peer_group_name"]
    .dropna()
    .unique()
)

for group in groups:

    temp = df[
        df["peer_group_name"] == group
    ].copy()

    temp.to_excel(
        writer,
        sheet_name=str(group)[:31],
        index=False,
    )

writer.close()

print()

print("Excel Created")

print(excel_file)
wb = load_workbook(excel_file)
green = PatternFill(
    fill_type="solid",
    start_color="90EE90",
)

yellow = PatternFill(
    fill_type="solid",
    start_color="FFF59D",
)

red = PatternFill(
    fill_type="solid",
    start_color="FF9999",
)

gold = PatternFill(
    fill_type="solid",
    start_color="FFD700",
)
for sheet in wb.sheetnames:

    ws = wb[sheet]

    headers = [
        cell.value
        for cell in ws[1]
    ]

    if "is_benchmark" not in headers:
        continue

    idx = headers.index(
        "is_benchmark"
    ) + 1

    for row in range(
        2,
        ws.max_row + 1,
    ):

        value = ws.cell(
            row=row,
            column=idx,
        ).value

        if value == 1:

            for col in range(
                1,
                ws.max_column + 1,
            ):

                ws.cell(
                    row=row,
                    column=col,
                ).fill = gold
for sheet in wb.sheetnames:

    ws = wb[sheet]

    last = ws.max_row + 2

    ws.cell(
        row=last,
        column=1,
    ).value = "Median Summary"
wb.save(excel_file)

conn.close()

print()

print("=" * 70)

print("DAY 20 COMPLETED")

print("=" * 70)                    